from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string, get_column_letter

# ====== CONFIGURAÇÕES ======
CAMINHO_ORIGEM = r"C:\Users\Ronny Santos\Desktop\Pasta2.xlsx"
CAMINHO_DESTINO = r"C:\Users\Ronny Santos\Desktop\resultado_final.xlsx"
CABEÇALHOS = [
    "Data Atendimento", "Cliente", "Início", "Término",
    "Tempo de Execução", "SLA", "Motivo Visita", "Reparo Realizado"
]
COLUNAS_CENTRALIZAR = ['A', 'C', 'D', 'E', 'F']

# ====== FUNÇÕES ======

def ajustar_largura_colunas(ws):
    """Ajusta automaticamente a largura das colunas com base no maior conteúdo."""
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

def centralizar_colunas(ws, colunas):
    """Centraliza o conteúdo das colunas especificadas."""
    alignment = Alignment(horizontal='center', vertical='center')
    for col_letter in colunas:
        col_idx = column_index_from_string(col_letter)
        for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for cell in col_cells:
                cell.alignment = alignment

def extrair_dados(ws_origem, ws_novo):
    """Extrai dados da planilha original e adiciona na nova."""
    for row in ws_origem.iter_rows():
        for cell in row:
            if cell.value == "Cliente":
                linha = cell.row
                coluna = cell.column
                try:
                    data_atend = ws_origem.cell(row=linha + 3, column=coluna).value
                    cliente = ws_origem.cell(row=linha + 1, column=coluna).value
                    motivo_visita = ws_origem.cell(row=linha + 5, column=coluna).value
                    reparo = ws_origem.cell(row=linha + 5, column=23).value  # Coluna AB
                    inicio = ws_origem.cell(row=linha + 3, column=5).value
                    termino = ws_origem.cell(row=linha + 3, column=7).value
                    tempo_exec = ws_origem.cell(row=linha + 3, column=13).value
                    sla = ws_origem.cell(row=linha + 3, column=14).value

                    ws_novo.append([
                        data_atend, cliente, inicio, termino,
                        tempo_exec, sla, motivo_visita, reparo
                    ])

                except Exception as e:
                    print(f"⚠️ Erro ao extrair dados na linha {linha}: {e}")

# ====== EXECUÇÃO ======

# Abre a planilha original
wb_origem = load_workbook(CAMINHO_ORIGEM, data_only=True)
ws_origem = wb_origem.active

# Cria a nova planilha
wb_novo = Workbook()
ws_novo = wb_novo.active
ws_novo.title = "Resultado"

# Adiciona cabeçalhos
ws_novo.append(CABEÇALHOS)

# Extrai os dados
extrair_dados(ws_origem, ws_novo)

# Ajustes finais
centralizar_colunas(ws_novo, COLUNAS_CENTRALIZAR)
ajustar_largura_colunas(ws_novo)

# Congelar cabeçalho
ws_novo.freeze_panes = 'A2'

# Salva a nova planilha
wb_novo.save(CAMINHO_DESTINO)
print("✅ Dados extraídos e planilha salva com sucesso!")
